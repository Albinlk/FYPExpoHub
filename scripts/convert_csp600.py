import openpyxl, csv, re

def clean_supervisor(val):
    s = str(val).strip() if val else ''
    s = re.sub(r'^SV\s*-\s*', '', s).strip()
    return s

def clean_examiner(val):
    s = str(val).strip() if val else ''
    s = re.sub(r'^EX\s*-\s*', '', s).strip()
    return s

def parse_title(val):
    title_str = str(val).strip() if val else ''
    if title_str.startswith('=UPPER('):
        m = re.search(r'"([^"]+)"', title_str)
        if m:
            title_str = m.group(1)
    return title_str.upper()

rows = []

# === 1. Parse JADUAL CSP600 MAC 2026 V2.xlsx (has titles) ===
wb1 = openpyxl.load_workbook(r'C:\Users\l_alb\Downloads\JADUAL CSP600 MAC 2026 V2.xlsx', data_only=True)
for sheet_name in wb1.sheetnames:
    ws = wb1[sheet_name]
    for row in ws.iter_rows(min_row=5, values_only=True):
        if row[0] is not None and isinstance(row[0], (int, float)):
            rows.append({
                'title': parse_title(row[4]),
                'technology_tags': '',
                'supervisor_display_name': clean_supervisor(row[5]),
                'programme_code': 'CS600',
                'short_description': parse_title(row[4]),
                'category': 'Computer Science',
                'team_display_names': str(row[3]).strip().upper() if row[3] else '',
                'student_id': str(int(row[2])) if row[2] else '',
                'examiner': clean_examiner(row[6]),
                'session': sheet_name,
                'time_slot': str(row[1]).strip() if row[1] else '',
            })

# === 2. Parse Sesi pembentangan CSP600 - CS251 - 2 July.xlsx (no titles) ===
wb2 = openpyxl.load_workbook(r'C:\Users\l_alb\Downloads\Sesi pembentangan CSP600 - CS251 - 2 July.xlsx', data_only=True)
for sheet_name in wb2.sheetnames:
    ws = wb2[sheet_name]
    current_session = ''
    for row in ws.iter_rows(min_row=1, values_only=True):
        # Detect session headers like 'SESI 1 (DS 3B) 1.30-6pm'
        if row[0] and isinstance(row[0], str) and row[0].startswith('SESI'):
            current_session = row[0]
            continue
        # Skip header rows
        if row[0] == 'Bil':
            continue
        # Data rows have a numeric Bil
        if row[0] is not None and isinstance(row[0], (int, float)):
            student_name = str(row[2]).strip().upper() if row[2] else ''
            rows.append({
                'title': f'CS251 PROPOSAL - {student_name}',
                'technology_tags': '',
                'supervisor_display_name': clean_supervisor(row[3]),
                'programme_code': 'CS251',
                'short_description': f'CSP600 proposal by {student_name}',
                'category': 'Computer Science',
                'team_display_names': student_name,
                'student_id': str(int(row[1])) if row[1] else '',
                'examiner': clean_examiner(row[4]),
                'session': current_session or 'CS251 - 2 July',
                'time_slot': str(row[5]).strip() if row[5] else '',
            })

# Write combined CSV
output_path = r'D:\MobileAppDev\FYPExpoHub\assets\data\csp600-proposals.csv'
with open(output_path, 'w', newline='', encoding='utf-8') as f:
    writer = csv.DictWriter(f, fieldnames=[
        'title', 'technology_tags', 'supervisor_display_name',
        'programme_code', 'short_description', 'category',
        'team_display_names', 'student_id', 'examiner',
        'session', 'time_slot',
    ])
    writer.writeheader()
    writer.writerows(rows)

# Stats
cs600 = sum(1 for r in rows if r['programme_code'] == 'CS600')
cs251 = sum(1 for r in rows if r['programme_code'] == 'CS251')
print(f'Written {len(rows)} total proposals')
print(f'  CS600 (with titles): {cs600}')
print(f'  CS251 (no titles):   {cs251}')

# Print some CS251 entries
print('\nCS251 sample:')
for r in rows[cs600:cs600+5]:
    print(f"  {r['title'][:70]}")
    print(f"    SV: {r['supervisor_display_name']} | Session: {r['session']}")
