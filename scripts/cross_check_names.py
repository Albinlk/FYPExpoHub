import openpyxl
import re
from collections import OrderedDict

# ===== 1. Parse attendance file =====
att_wb = openpyxl.load_workbook(r'D:\Downloads\senarai_kedatangan_pelajar_by_group (5).xlsx', data_only=True)
ws = att_wb['Table 1']

attendance_names = {}  # matric_id -> name
for r in range(1, ws.max_row + 1):
    col2 = ws.cell(row=r, column=2).value
    if col2 is None:
        continue
    val = str(col2).strip()
    # Match pattern: "BIL  NO_PELAJAR  NAMA ..."
    # Lines look like: "1        2023680626     ADAM BIN MOHD AKIB"
    parts = val.split()
    if len(parts) >= 3:
        # Check if second token looks like a matric number (9-10 digits)
        if re.match(r'^\d{9,10}$', parts[1]):
            matric = parts[1]
            name = ' '.join(parts[2:]).strip()
            # Clean up the name (remove extra spaces, etc.)
            name = re.sub(r'\s+', ' ', name).strip()
            attendance_names[matric] = name

print(f"Total students in attendance file: {len(attendance_names)}")

# ===== 2. Parse project Excel file =====
proj_wb = openpyxl.load_workbook(r'D:\Downloads\(LATEST) CSP650-DewanSeminar-ListName-Layout (5).xlsx', data_only=True)

sheet_map = OrderedDict([
    ('CS230 - List Name', 'CS230'),
    ('CS251 - List Name', 'CS251'),
    ('CS253 - List Name', 'CS253'),
    ('LATEST CS255 - List Name', 'CS255'),
    ('CS266 - List Name', 'CS266'),
])

project_names = {}  # matric_id -> {name, course, group, title}
for sheet_name, course in sheet_map.items():
    ws = proj_wb[sheet_name]
    for row in ws.iter_rows(min_row=4, values_only=True):
        vals = [v for v in row]
        if len(vals) < 8:
            continue
        no = vals[0]
        booth = vals[1]
        group = str(vals[2]).strip() if vals[2] else ''
        matric_raw = vals[3]
        student_name = str(vals[4]).strip() if vals[4] else ''
        title = str(vals[5]).strip() if vals[5] else ''
        
        # Skip header rows
        try:
            int(float(str(no)))
        except (ValueError, TypeError):
            continue
        
        if not title or title == 'None' or not student_name or student_name == 'None':
            continue
        
        # Clean matric
        matric = str(matric_raw).replace('.0', '').strip()
        name_clean = re.sub(r'\s+', ' ', student_name).strip()
        
        project_names[matric] = {
            'name': name_clean,
            'course': course,
            'group': group,
        }

print(f"Total students in project Excel: {len(project_names)}")

# ===== 3. Cross-check =====
att_matrics = set(attendance_names.keys())
proj_matrics = set(project_names.keys())

matched = att_matrics & proj_matrics
only_attendance = att_matrics - proj_matrics
only_project = proj_matrics - att_matrics

print(f"\n=== CROSS-CHECK RESULTS ===")
print(f"Matched (in both files): {len(matched)}")
print(f"Only in attendance file: {len(only_attendance)}")
print(f"Only in project Excel: {len(only_project)}")

if only_attendance:
    print(f"\n--- Students in attendance file but NOT in project Excel ({len(only_attendance)}) ---")
    for m in sorted(only_attendance):
        print(f"  {m}  {attendance_names[m][:50]}")

if only_project:
    print(f"\n--- Students in project Excel but NOT in attendance file ({len(only_project)}) ---")
    for m in sorted(only_project):
        info = project_names[m]
        print(f"  {m}  {info['name'][:50]}  ({info['course']} / {info['group']})")

# Name match quality check
name_mismatches = []
for m in matched:
    att_name = attendance_names[m].upper()
    proj_name = project_names[m]['name'].upper()
    if att_name != proj_name:
        name_mismatches.append((m, attendance_names[m], project_names[m]['name']))

if name_mismatches:
    print(f"\n--- Name mismatches (same matric, different name spelling) ---")
    for m, a, p in name_mismatches:
        print(f"  {m}")
        print(f"    Attendance: {a}")
        print(f"    Project:    {p}")
