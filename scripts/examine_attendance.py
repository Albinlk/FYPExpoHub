import openpyxl

wb = openpyxl.load_workbook(r'D:\Downloads\senarai_kedatangan_pelajar_by_group (5).xlsx', data_only=True)
print(f'Sheets: {wb.sheetnames}')
for name in wb.sheetnames:
    ws = wb[name]
    print(f'\n=== {name} (rows={ws.max_row}, cols={ws.max_column}) ===')
    for r, row in enumerate(ws.iter_rows(min_row=1, max_row=min(8, ws.max_row), values_only=True), 1):
        vals = [str(v)[:50] if v is not None else '' for v in row]
        print(f'  Row {r}: {vals}')
