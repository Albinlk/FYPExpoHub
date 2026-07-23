import openpyxl
import os
import math
from PIL import Image, ImageDraw, ImageFont

EXCEL_PATH = r'D:\Downloads\(LATEST) CSP650-DewanSeminar-ListName-Layout (5).xlsx'
OUTPUT_DIR = r'D:\MobileAppDev\FYPExpoHub\web\project_images'

COURSE_PALETTE = {
    'CS230': {
        'primary': '#0A1628', 'secondary': '#1B3A6B',
        'accent': '#F0C929', 'accent2': '#F4D03F',
        'icon': '💻',
    },
    'CS251': {
        'primary': '#0A1F14', 'secondary': '#1B4A33',
        'accent': '#2ECC71', 'accent2': '#58D68D',
        'icon': '🌐',
    },
    'CS253': {
        'primary': '#1A0A0A', 'secondary': '#4A1B1B',
        'accent': '#E74C3C', 'accent2': '#EC7063',
        'icon': '🛡️',
    },
    'CS255': {
        'primary': '#0A1628', 'secondary': '#1B3A6B',
        'accent': '#3498DB', 'accent2': '#5DADE2',
        'icon': '🔒',
    },
    'CS266': {
        'primary': '#140A1F', 'secondary': '#2D1B4A',
        'accent': '#9B59B6', 'accent2': '#AF7AC5',
        'icon': '📱',
    },
}

COURSE_SHEET_NAMES = {
    'CS230': 'CS230 - List Name',
    'CS251': 'CS251 - List Name',
    'CS253': 'CS253 - List Name',
    'CS255': 'LATEST CS255 - List Name',
    'CS266': 'CS266 - List Name',
}

W, H = 800, 450

def try_font(size, bold=False):
    candidates = [
        ("C:\\Windows\\Fonts\\segoeuib.ttf" if bold else "C:\\Windows\\Fonts\\segoeui.ttf"),
        ("C:\\Windows\\Fonts\\arialbd.ttf" if bold else "C:\\Windows\\Fonts\\arial.ttf"),
        ("C:\\Windows\\Fonts\\calibrib.ttf" if bold else "C:\\Windows\\Fonts\\calibri.ttf"),
    ]
    for fp in candidates:
        if os.path.exists(fp):
            try:
                return ImageFont.truetype(fp, size)
            except:
                pass
    return ImageFont.load_default()

def hexrgb(h):
    h = h.lstrip('#')
    return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))

def draw_diagonal_split(draw, c1, c2):
    """Upper-left to bottom-right diagonal two-tone."""
    pts = [(W, 0), (W, H), (0, H), (0, 0)]
    c = hexrgb(c1)
    draw.polygon(pts, fill=c)
    # overlay a triangle
    t = [(0, 0), (W, 0), (0, H)]
    draw.polygon(t, fill=hexrgb(c2))

def draw_circle_pattern(draw, accent, accent2):
    ac = hexrgb(accent)
    ac2 = hexrgb(accent2)
    # Large soft circle bottom-right
    draw.ellipse([400, 150, 750, 500], fill=(*ac, 30) if has_transparency else ac, outline=None)
    # Small circle top-right
    draw.ellipse([620, -60, 820, 140], fill=(*ac2, 20) if has_transparency else ac2)
    # Thin ring
    draw.ellipse([580, 80, 780, 280], outline=ac2, width=3)

    # Grid dots pattern
    for x in range(20, W, 40):
        for y in range(20, H, 40):
            if (x // 40 + y // 40) % 3 == 0:
                draw.ellipse([x-1, y-1, x+1, y+1], fill=(255, 255, 255, 20))

    # Accent bar top
    draw.rectangle([0, 0, W, 5], fill=ac)

# Check if RGBA is supported
has_transparency = True
try:
    test = Image.new('RGBA', (10, 10), (0, 0, 0, 0))
except:
    has_transparency = False

def wrap_text(text, font, max_w, draw):
    words = text.split()
    lines = []
    cur = ''
    for w in words:
        test = (cur + ' ' + w).strip()
        bb = draw.textbbox((0, 0), test, font=font)
        tw = bb[2] - bb[0]
        if tw <= max_w:
            cur = test
        else:
            if cur:
                lines.append(cur)
            cur = w
    if cur:
        lines.append(cur)
    return lines

def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    f_title = try_font(40, bold=True)
    f_subtitle = try_font(22)
    f_small = try_font(17)
    f_tiny = try_font(13)

    total = 0
    global_idx = 0

    for course_code in ['CS230', 'CS251', 'CS253', 'CS255', 'CS266']:
        ws = wb[COURSE_SHEET_NAMES[course_code]]
        pal = COURSE_PALETTE[course_code]
        max_row = ws.max_row

        for r in range(1, max_row + 1):
            a = ws.cell(row=r, column=1).value
            if a is None:
                continue
            try:
                int(float(a))
            except (ValueError, TypeError):
                continue

            title = ws.cell(row=r, column=6).value
            student = ws.cell(row=r, column=5).value
            group = ws.cell(row=r, column=3).value

            if not title:
                continue

            global_idx += 1
            pid = f'proj-{course_code.lower()}-{global_idx:03d}'

            title_str = str(title).strip()
            student_str = str(student).strip() if student else ''
            group_str = str(group).strip() if group else course_code

            img = Image.new('RGBA', (W, H), (255, 255, 255, 255))
            draw = ImageDraw.Draw(img)

            # Background diagonal split
            c1 = pal['primary']
            c2 = pal['secondary']
            draw.polygon([(0, 0), (W, 0), (0, H)], fill=hexrgb(c2))
            draw.polygon([(W, 0), (W, H), (0, H)], fill=hexrgb(c1))

            # Large translucent circle
            ac = hexrgb(pal['accent'])
            draw.ellipse([380, 120, 780, 520], fill=(ac[0], ac[1], ac[2], 25))

            # Small circle top-right
            ac2 = hexrgb(pal['accent2'])
            draw.ellipse([650, -80, 850, 120], fill=(ac2[0], ac2[1], ac2[2], 30))

            # Accent bar
            draw.rectangle([0, 0, W, 4], fill=ac)

            # Diagonal line accent
            draw.line([(200, 0), (W, 280)], fill=ac, width=2)
            draw.line([(250, 0), (W, 300)], fill=ac2, width=1)

            # Dots grid
            for x in range(30, W, 50):
                for y in range(30, H, 50):
                    if ((x // 50) + (y // 50)) % 3 == 0:
                        dc = (255, 255, 255, 15)
                        draw.ellipse([x-1.5, y-1.5, x+1.5, y+1.5], fill=dc)

            # Course badge top-left
            badge_text = f'{course_code}  •  {group_str}'
            bb = draw.textbbox((0, 0), badge_text, font=f_small)
            bw = bb[2] - bb[0] + 24
            bh = bb[3] - bb[1] + 12
            draw.rounded_rectangle([25, 20, 25 + bw, 20 + bh], radius=6, fill=ac)
            draw.text((25 + 12, 20 + 6), badge_text, fill='white', font=f_small)

            # Project title (left-aligned, bottom section)
            title_lines = wrap_text(title_str, f_title, W - 100, draw)
            y_start = 120
            for i, line in enumerate(title_lines[:4]):
                # Subtle shadow
                draw.text((42, y_start + 2), line, fill=(0, 0, 0, 60), font=f_title)
                draw.text((40, y_start), line, fill='white', font=f_title)
                y_start += 50

            # Bottom accent line
            draw.rectangle([40, 310, 200, 313], fill=ac)

            # Student name
            if student_str:
                draw.text((40, 340), student_str, fill=(200, 200, 200), font=f_subtitle)

            out_path = os.path.join(OUTPUT_DIR, f'{pid}.png')
            # Convert to RGB for saving
            rgb_img = Image.new('RGB', (W, H), (255, 255, 255))
            rgb_img.paste(img, (0, 0), img)
            rgb_img.save(out_path, quality=92, optimize=True)
            total += 1

            if total % 50 == 0:
                print(f'  Generated {total} images...')

    print(f'\nDone! Generated {total} project images.')
    print(f'Output: {OUTPUT_DIR}')

if __name__ == '__main__':
    main()
