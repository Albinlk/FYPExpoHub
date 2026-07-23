import openpyxl
import re

wb = openpyxl.load_workbook(r'D:\Downloads\senarai_kedatangan_pelajar_by_group (5).xlsx', data_only=True)
ws = wb['Table 1']

# Print all non-empty rows to understand the structure
for r in range(1, ws.max_row + 1):
    row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
    non_empty = [(c, v) for c, v in enumerate(row, 1) if v is not None]
    if non_empty:
        print(f'Row {r}: {non_empty[:8]}')
