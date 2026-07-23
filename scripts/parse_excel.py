import openpyxl
import re

EXCEL_PATH = r'D:\Downloads\(LATEST) CSP650-DewanSeminar-ListName-Layout (2).xlsx'
OUTPUT_PATH = r'D:\MobileAppDev\FYPExpoHub\lib\core\data\excel_data.dart'

COURSE_CATEGORY = {
    'CS230': 'Computer Science',
    'CS251': 'Networking & Communication',
    'CS253': 'Cybersecurity',
    'CS255': 'Network Security & Infrastructure',
    'CS266': 'Software Engineering & Applications',
}

COURSE_PROGRAMME_NAME = {
    'CS230': 'CS230 (Computer Science)',
    'CS251': 'CS251 (Networking & Communication)',
    'CS253': 'CS253 (Cybersecurity)',
    'CS255': 'CS255 (Network Security & Infrastructure)',
    'CS266': 'CS266 (Software Engineering & Applications)',
}

COURSE_SHEET_NAMES = {
    'CS230': 'CS230 - List Name',
    'CS251': 'CS251 - List Name',
    'CS253': 'CS253 - List Name',
    'CS255': 'LATEST CS255 - List Name',
    'CS266': 'CS266 - List Name',
}

def parse_date_compat(date_str):
    d = str(date_str).upper().strip()
    day = 6
    if '07' in d:
        day = 7
    return f'DateTime(2026, 08, {day:02d})'

def extract_zone(booth_no):
    if not booth_no:
        return 'General', ''
    parts = str(booth_no).strip().split('-')
    if len(parts) >= 2:
        return parts[0], '-'.join(parts[1:])
    return 'General', booth_no

def make_slug(title):
    slug = title.lower().strip()
    slug = re.sub(r'[^a-z0-9\s-]', '', slug)
    slug = re.sub(r'[\s-]+', '-', slug)
    return slug[:80]

def esc(s):
    """Escape a string for Dart single-quoted string literal."""
    if s is None:
        return 'null'
    s = str(s)
    s = s.replace('\\', '\\\\')
    s = s.replace("'", "\\'")
    s = s.replace('\n', '\\n')
    s = s.replace('\r', '')
    return f"'{s}'"

def esc_nullable(s):
    if s is None or str(s).strip() == '':
        return 'null'
    return esc(str(s).strip())

def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    all_projects = []
    all_booths = []
    all_schedule = []
    seen_booths = set()

    schedule_index = [0]
    project_index = [0]

    for course_code in ['CS230', 'CS251', 'CS253', 'CS255', 'CS266']:
        sheet_name = COURSE_SHEET_NAMES[course_code]
        ws = wb[sheet_name]
        max_row = ws.max_row
        print(f'Processing {course_code} ({sheet_name}): {max_row} rows')

        category = COURSE_CATEGORY[course_code]
        prog_name = COURSE_PROGRAMME_NAME[course_code]

        # Collect all session blocks and data rows
        sessions = []  # list of (location, date_expr, data_start_row, data_end_row)
        standalone_rows = []  # rows with project data but no session header

        # First, collect all LOCATION/DATE headers and their approximate ranges
        location_info = {}  # row_num -> (location, date_expr)
        for r in range(1, max_row + 1):
            a = ws.cell(row=r, column=1).value
            if a and isinstance(a, str):
                if 'LOCATION' in a:
                    location = a.split(':')[-1].strip()
                    date_expr = ''
                    if r + 1 <= max_row:
                        date_str = ws.cell(row=r+1, column=1).value or ''
                        date_expr = parse_date_compat(str(date_str))
                    location_info[r] = (location, date_expr)

        # Now collect all data rows (rows with numeric No.)
        # Group them into session blocks based on proximity to LOCATION headers
        location_rows = sorted(location_info.keys())
        
        # For finding which session a data row belongs to
        def find_session(row_num):
            prev_loc_row = 0
            prev_loc = ('', 'DateTime(2026, 08, 06)')
            for lr in location_rows:
                if lr < row_num:
                    prev_loc_row = lr
                    prev_loc = location_info[lr]
                else:
                    break
            return prev_loc

        # Scan all rows for project data
        current_session = ('', 'DateTime(2026, 08, 06)')
        current_session_start = 0
        
        for r in range(1, max_row + 1):
            a = ws.cell(row=r, column=1).value
            
            # Track session changes
            if r in location_info:
                current_session = location_info[r]
                current_session_start = r
            
            # Check if this is a data row (has numeric No.)
            if a is None:
                continue
            try:
                no_val = int(float(a))
            except (ValueError, TypeError):
                continue
            
            row_data = [ws.cell(row=r, column=c).value for c in range(1, 9)]
            student_name = str(row_data[4]).strip() if row_data[4] else ''
            title = str(row_data[5]).strip() if row_data[5] else ''
            
            if not title or not student_name:
                continue
            
            booth_no = str(row_data[1]).strip() if row_data[1] else ''
            group = str(row_data[2]).strip() if row_data[2] else ''
            raw_matric = row_data[3]
            supervisor = str(row_data[6]).strip() if row_data[6] else ''
            examiner = str(row_data[7]).strip() if row_data[7] else ''

            matric_id = ''
            if raw_matric is not None:
                if isinstance(raw_matric, (int, float)):
                    matric_id = str(int(raw_matric))
                else:
                    matric_id = str(raw_matric).strip()

            project_index[0] += 1
            pid = f'proj-{course_code.lower()}-{project_index[0]:03d}'
            slug = make_slug(title)
            if not slug:
                slug = f'project-{project_index[0]}'
            
            short_desc = f'Final Year Project - {category}'
            tags = '["FYP"]'
            if 'AI' in title or 'Machine Learning' in title or 'Deep Learning' in title or 'Intelligence' in title:
                tags = '["AI", "Machine Learning"]'
            elif 'IoT' in title or 'Smart' in title or 'Sensor' in title:
                tags = '["IoT", "Embedded Systems"]'
            elif 'Web' in title or 'Mobile' in title or 'App' in title:
                tags = '["Web", "Mobile"]'
            elif 'Security' in title or 'Attack' in title or 'Threat' in title or 'Penetration' in title or 'Honeypot' in title or 'IDS' in title:
                tags = '["Security", "Network"]'
            elif 'Blockchain' in title or 'QR' in title:
                tags = '["Blockchain", "Security"]'

            booth_id = 'null'
            booth_number = 'null'
            booth_zone = 'null'
            if booth_no:
                zone, _ = extract_zone(booth_no)
                booth_id = esc(f'booth-{booth_no}')
                booth_number = esc(booth_no)
                booth_zone = esc(zone)
                
                if booth_no not in seen_booths:
                    seen_booths.add(booth_no)
                    all_booths.append({
                        'id': esc(f'booth-{booth_no}'),
                        'event_id': esc('fskm-fyp-2026'),
                        'booth_number': esc(booth_no),
                        'zone': esc(zone),
                        'location_note': esc(current_session[0]),
                        'static_floor_plan_url': 'null',
                        'project_id': esc(pid),
                        'publication_status': esc('published'),
                        'created_at': 'DateTime(2026, 07, 01)',
                        'updated_at': 'DateTime(2026, 07, 01)',
                        'published_at': 'DateTime(2026, 07, 01)',
                    })
                else:
                    # Update project_id reference for existing booth
                    for b in all_booths:
                        if b['booth_number'] == esc(booth_no):
                            b['project_id'] = esc(pid)
                            break

            all_projects.append({
                'id': esc(pid),
                'event_id': esc('fskm-fyp-2026'),
                'slug': esc(slug),
                'title': esc(title),
                'matric_id': esc(matric_id) if matric_id else 'null',
                'programme_code': esc(group) if group else esc(course_code),
                'programme_name': esc(prog_name),
                'short_description': esc(short_desc),
                'category': esc(category),
                'technology_tags': tags,
                'booth_id': booth_id,
                'booth_number': booth_number,
                'booth_zone': booth_zone,
                'cover_image_url': esc('assets/images/project_placeholder.jpg'),
                'poster_url': 'null',
                'team_display_names': f'[{esc(student_name)}]',
                'supervisor_display_name': esc(supervisor) if supervisor else esc('TBD'),
                'examiner_display_name': esc_nullable(examiner),
                'demo_url': 'null',
                'video_url': 'null',
                'repository_url': 'null',
                'featured': 'false',
                'publication_status': esc('published'),
                'created_at': 'DateTime(2026, 07, 01)',
                'updated_at': 'DateTime(2026, 07, 01)',
                'published_at': 'DateTime(2026, 07, 01)',
            })

        # Create schedule items from location sessions
        for loc_row, (location, date_expr) in location_info.items():
            schedule_index[0] += 1
            session_label = 'Session'
            if 'DS5' in location.upper():
                session_label = 'DS5'
            elif 'DS6' in location.upper():
                session_label = 'DS6'
            elif 'DS7' in location.upper():
                session_label = 'DS7'
            elif 'DS8' in location.upper():
                session_label = 'DS8'
            elif 'BK' in location.upper():
                session_label = location.strip().replace(' ', '_')
            
            date_str_raw = ws.cell(row=loc_row+1, column=1).value or ''
            day_label = 'Day 1'
            if '07' in str(date_str_raw):
                day_label = 'Day 2'

            all_schedule.append({
                'id': esc(f'sch-{course_code.lower()}-{schedule_index[0]:02d}'),
                'event_id': esc('fskm-fyp-2026'),
                'title': esc(f'{course_code} Presentation - {session_label} ({day_label})'),
                'venue': esc(location),
                'date': date_expr,
                'start_at': esc('09:00'),
                'end_at': esc('17:00'),
                'audience': esc(course_code),
                'description': esc(f'{course_code} final year project presentation session at {location}'),
                'visibility': esc('public'),
                'publication_status': esc('published'),
                'source_import_id': 'null',
                'source_staging_id': 'null',
                'created_at': 'DateTime(2026, 07, 01)',
                'updated_at': 'DateTime(2026, 07, 01)',
                'published_at': 'DateTime(2026, 07, 01)',
            })

        print(f'  -> {project_index[0] - len(all_projects) + len([p for p in all_projects if p["id"].startswith(esc(f"proj-{course_code.lower()}"))])} projects, {sum(1 for p in all_projects if p["id"].startswith(f"\'proj-{course_code.lower()}\'"))} total in this course')

    # Write output
    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        f.write('// Auto-generated from Excel parser. Do not edit manually.\n')
        f.write('// Generated from: CSP650-DewanSeminar-ListName-Layout (2).xlsx\n\n')
        f.write('class ExcelData {\n')
        f.write('  static const String eventId = "fskm-fyp-2026";\n\n')

        # Projects
        f.write(f'  static const int projectCount = {len(all_projects)};\n')
        f.write('  static List<Map<String, dynamic>> get allProjects => [\n')
        for p in all_projects:
            f.write('    {\n')
            for k, v in p.items():
                f.write(f'      "{k}": {v},\n')
            f.write('    },\n')
        f.write('  ];\n\n')

        # Booths
        f.write(f'  static const int boothCount = {len(all_booths)};\n')
        f.write('  static List<Map<String, dynamic>> get allBooths => [\n')
        for b in all_booths:
            f.write('    {\n')
            for k, v in b.items():
                f.write(f'      "{k}": {v},\n')
            f.write('    },\n')
        f.write('  ];\n\n')

        # Schedule
        f.write(f'  static const int scheduleCount = {len(all_schedule)};\n')
        f.write('  static List<Map<String, dynamic>> get allScheduleItems => [\n')
        for s in all_schedule:
            f.write('    {\n')
            for k, v in s.items():
                f.write(f'      "{k}": {v},\n')
            f.write('    },\n')
        f.write('  ];\n')
        f.write('}\n')

    total_projects_final = sum(1 for p in all_projects) 
    print(f'\nDone! Generated:')
    print(f'  Projects: {total_projects_final}')
    print(f'  Booths: {len(all_booths)}')
    print(f'  ScheduleItems: {len(all_schedule)}')
    print(f'  Output: {OUTPUT_PATH}')

if __name__ == '__main__':
    main()
